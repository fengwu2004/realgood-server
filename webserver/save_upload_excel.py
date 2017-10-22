from io import BytesIO
from data import storemgr
from data.suggest import Suggest, Consultor
from openpyxl import load_workbook
from stock.consultor_manager import ConsultorManager
from webserver.RequestBaseManager import RequestBaseManager
from datetime import datetime

def getItems(ws) -> set:
    
    results = set()
    
    for i in range(1, ws.max_row):
    
        obj = Suggest()
        
        index = str(i + 1)
        
        temp = str(ws['A' + index].value)
        
        obj.date = datetime.strptime(temp[0:10], '%Y-%m-%d').strftime('%Y-%m-%d')

        obj.stockId = ws['D' + index].value

        obj.stockName = storemgr.getStockName(obj.stockId)
        
        if obj.stockName is None:
            
            continue

        consultorName = ws['C' + index].value

        consultorCompany = ws['B' + index].value

        consultor = ConsultorManager.instance().retriveConsultor(consultorName, consultorCompany)

        obj.consultor = consultor

        obj.consultorId = consultor.id

        results.add(obj)
        
    return results

class SaveSuggestExcel(RequestBaseManager):
    
    def post (self, *args, **kwargs):
        
        data = self.request.files['upload'][0].body

        wb = load_workbook(filename = BytesIO(data))
        
        ws = wb.active
        
        storemgr.saveSuggests(getItems(ws))

        self.write({'success': 1})