import json
from data.suggest import SuggestInfo
from openpyxl import Workbook
from analyse import calc_detail_suggest
from webserver import tokenManager
from webserver.RequestBaseManager import RequestBaseManager

def createExcel(items:[SuggestInfo]):
    
    wb = Workbook()

    ws = wb.active
    
    index = 1
    
    for item in items:
        
        ws['A' + str(index)] = item.suggeststock.stockName

        ws['B' + str(index)] = item.suggeststock.name

        ws['C' + str(index)] = item.suggeststock.date

        ws['D' + str(index)] = item.suggeststock.company

        ws['E' + str(index)] = item.counts[0]

        ws['F' + str(index)] = item.counts[1]

        ws['G' + str(index)] = item.counts[2]

        ws['H' + str(index)] = item.counts[3]

        ws['I' + str(index)] = item.trends[0]
        
        ws['J' + str(index)] = item.trends[1]

        ws['K' + str(index)] = item.trends[2]

        ws['L' + str(index)] = item.trends[3]

        ws['M' + str(index)] = item.trends[4]

        index += 1

    wb.save('/Users/yan/Desktop/ccc/asdlsk.xlsx')

class FindHistorySuggestWithTrends(RequestBaseManager):
    
    def post (self, *args, **kwargs):
        
        # data = json.loads(self.request.body.decode('utf-8'))
        #
        # if not 'token' in data or not tokenManager.TokenManagerInstance().checkToken(data['token']):
        #
        #     self.write({'success': -1})
        #
        #     return
        
        results = calc_detail_suggest.run()

        createExcel(results)
            
        self.write({'success': 1})