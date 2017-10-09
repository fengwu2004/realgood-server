from openpyxl import Workbook, load_workbook
from typing import List, Tuple, Dict
from datetime import datetime
from analyse import SuggestHistoryManager
from data.stock_info import ConsultorWinsLevel, Consultor, SuggestStock, SuggestScore


class RateSetting(object):

    def __init__(self, start, end, score):

        self.start = start

        self.end = end

        self.score = score

_instance = None

class ConsultorWinsLevelManager(object):

    @classmethod
    def instance(cls):

        global _instance

        if _instance is None:

            _instance = ConsultorWinsLevelManager()

        return _instance

    def __init__(self):

        self.consultors_wins_level = dict()

    def addConsultorScore(self, score:int, suggest:SuggestStock):

        suggestscore = SuggestScore(score, suggest)

        if suggest.consultor not in self.consultors_wins_level:

            self.consultors_wins_level[suggest.consultor] = [suggestscore]

        else:

            self.consultors_wins_level[suggest.consultor].append(suggestscore)

    def saveToDB(self):

        pass

    def loadFromDB(self):

        pass

_ratesettings = []

def initRateSetting() -> List[RateSetting]:

    global _ratesettings

    if len(_ratesettings) != 0:

        return _ratesettings

    wb = load_workbook('./ratesetting.xlsx')

    ws = wb.active

    for i in range(2, ws.max_row + 1):

        rs = RateSetting(ws['A' + str(i)].value, ws['B' + str(i)].value, ws['C' + str(i)].value)

        _ratesettings.append(rs)

    return _ratesettings

def retriveConsulterRate(v0:float, vMax:float) -> int:

    vMaxPercent = (vMax - v0) * 100 / v0

    items = initRateSetting()

    for item in items:

        if item.start <= vMaxPercent <= item.end:

            return item.score

    return 0

def getDayValues(stockId:str, dt0:datetime, dt1:datetime):

    stock = SuggestHistoryManager.instance().getStock(stockId)

    if stock is None:

        return None

    resulsts = []

    for dayvalue in stock.dayvalues:

        dt = datetime.strptime(dayvalue.date, '%Y/%m/%d')

        if dt < dt0:

            continue

        if dt > dt1:

            break

        resulsts.append(dayvalue)

    return resulsts

def getMaxIncrease(stockId:str, dt0:datetime, dt1:datetime) -> Tuple[float, float, int]:

    dayvalues = getDayValues(stockId, dt0, dt1)

    if dayvalues is None or len(dayvalues) == 1:

        return None

    v0 = dayvalues[0].close

    vMax = dayvalues[0].close

    vMaxDay = 0

    index = 0

    for dayvalue in dayvalues:

        index += 1

        if dayvalue.close > vMax:

            vMax = dayvalue.close

            vMaxDay = index

    return v0, vMax, vMaxDay



# print(retriveConsulterRate(100, 110))