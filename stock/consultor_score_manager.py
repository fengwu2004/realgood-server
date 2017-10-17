from typing import List
from collections import defaultdict
from data.suggest import SuggestScore, Suggest, Consultor
from openpyxl import load_workbook
from stock.consultor_manager import ConsultorManager

from stock.analyse_setting_manager import AnalyseSettingManager

_ratesettings = []

_instance = None

class RateSetting(object):

    def __init__(self, start, end, score):

        self.start = start

        self.end = end

        self.score = score

def initRateSetting() -> List[RateSetting]:

    global _ratesettings

    if len(_ratesettings) != 0:

        return _ratesettings

    wb = load_workbook('./stock/ratesetting.xlsx')

    ws = wb.active

    for i in range(2, ws.max_row + 1):

        rs = RateSetting(ws['A' + str(i)].value, ws['B' + str(i)].value, ws['C' + str(i)].value)

        _ratesettings.append(rs)

    return _ratesettings

def retriveConsulterRate(v0:float, vMax:float) -> int:

    vMaxPercent = (vMax - v0) * 100 / v0

    items = initRateSetting()

    for item in items:

        if item.start <= vMaxPercent <= item.end:

            return item.score

    return 0

class ConsultorScoreManager(object):

    @classmethod
    def instance(cls):

        global _instance

        if _instance is None:

            _instance = ConsultorScoreManager()

        return _instance

    def __init__(self):

        self.consultorscores = defaultdict(list)

    def addConsultorScore(self, score: int, suggest: Suggest):

        suggestscore = SuggestScore(score, suggest)

        self.consultorscores[suggest.consultorId].append(suggestscore)

    def getScores(self, consultorId: int) -> [SuggestScore]:

        return self.consultorscores[consultorId]

    def show(self):

        temps = list(map(lambda consultorId:(consultorId, self.getConsultorWeight(consultorId)), self.consultorscores))

        temps.sort(key = lambda item:item[1])

        for item in temps:

            consultor = ConsultorManager.instance().retriveConsultorBy(item[0])

            print(consultor.name, consultor.company, item[1])

    def getConsultorWeight(self, consultorId: int) -> int:

        scores = self.getScores(consultorId)

        if len(scores) == 0:

            return 0

        total = sum(list(map(lambda x: x.score, scores)))

        # return total

        return total / len(scores)

    def saveToDB(self):

        pass

    def loadFromDB(self):

        pass

# initRateSetting()